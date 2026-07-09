import streamlit as st
import pandas as pd
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO
import datetime

st.title("Crystal Report")

sponsor_file = st.file_uploader("Upload Data Sponsor (.xlsx)", type=["xlsx"])
batch_file   = st.file_uploader("Upload Batch Daily Report (.xls)", type=["xls"])

FMT_DATE    = "DD/MM/YYYY"
FMT_RP_MAIN = r'[$Rp-421]#,##0.00_);\([$Rp-421]#,##0.00\)'
FMT_RP_SUM  = r'[$Rp-409]#,##0.00_);\([$Rp-409]#,##0.00\)'
FMT_RP_MAP  = {165: FMT_RP_MAIN, 166: FMT_RP_SUM}

def load_sponsor_lookup(file):
    xls = pd.ExcelFile(file)
    df  = None
    for sheet in xls.sheet_names:
        tmp = pd.read_excel(xls, sheet_name=sheet)
        tmp.columns = tmp.columns.astype(str).str.strip()
        cols_up = [c.upper() for c in tmp.columns]
        if "ID_PARTNER" in cols_up and "TYPE" in cols_up and "RM" in cols_up:
            df = tmp; break
    if df is None:
        st.error("Kolom ID_Partner / Type / RM tidak ditemukan.")
        st.stop()
    df.columns = [c.strip().upper() for c in df.columns]
    df = df[["ID_PARTNER","TYPE","RM"]].copy()
    def cid(x):
        if pd.isna(x): return None
        try:   return str(int(float(x))).strip()
        except: return str(x).strip()
    df["ID_PARTNER"] = df["ID_PARTNER"].apply(cid)
    df["TYPE"] = df["TYPE"].apply(lambda x: "" if pd.isna(x) else str(x).strip())
    df["RM"]   = df["RM"].apply(lambda x: "" if pd.isna(x) else str(x).strip())
    df = df.sort_values(["TYPE","RM"], ascending=False).drop_duplicates(subset="ID_PARTNER", keep="first")
    return df.set_index("ID_PARTNER")[["TYPE","RM"]].to_dict(orient="index")

def clean_pid(x):
    if x is None or x == "": return None
    try:   return str(int(float(x))).strip()
    except: return str(x).strip()

def xlrd_val(wb_in, ws_in, r, c):
    cell = ws_in.cell(r, c)
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None, False
    if cell.ctype == xlrd.XL_CELL_DATE:
        tup = xlrd.xldate_as_tuple(cell.value, wb_in.datemode)
        return datetime.datetime(*tup), True
    return cell.value, False

def cell_fmt(wb_in, ws_in, r, c):
    xfi = ws_in.cell_xf_index(r, c)
    xf  = wb_in.xf_list[xfi]
    fk  = xf.format_key
    if fk in FMT_RP_MAP: return FMT_RP_MAP[fk]
    if fk == 164: return FMT_DATE
    return None

def process(file_bytes, lookup):
    wb_in = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    ws_in = wb_in.sheet_by_index(0)
    nrows = ws_in.nrows
    ncols = ws_in.ncols

    # Pass 1: build desig→type→amount, mark dup rows
    batch_desig_list = []
    dup_rows = set()
    cur_desig = None
    seen_pids = {}
    in_batch  = False

    for r in range(nrows):
        row = ws_in.row_values(r)
        if str(row[2]).strip() == 'PartnerID':
            if cur_desig is not None:
                batch_desig_list.append(cur_desig)
            cur_desig = defaultdict(lambda: defaultdict(float))
            seen_pids = {}
            in_batch  = True
            continue
        if str(row[0]).strip() == 'Total Batch Amount':
            in_batch = False; continue
        if not in_batch: continue
        pid = clean_pid(row[0])
        if pid is None:
            in_batch = False; continue
        if pid in seen_pids:
            dup_rows.add(r)
        else:
            seen_pids[pid] = r
        amount      = row[4] if isinstance(row[4], (int, float)) else 0
        designation = str(row[5]).strip() if row[5] != '' else ''
        if pid in lookup and designation:
            dt = lookup[pid]['TYPE']
            if dt in ('Mass','Middle','Major'):
                cur_desig[designation][dt] += amount
    if cur_desig is not None:
        batch_desig_list.append(cur_desig)

    # Pass 2: write to xlsx
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = wb_in.sheet_names()[0]

    # Column widths from template
    for c0 in range(ncols):
        ci = ws_in.colinfo_map.get(c0)
        w  = (ci.width / 256.0) if ci else 10
        ws_out.column_dimensions[get_column_letter(c0+1)].width = max(w, 6)
    # Mass/Middle/Major extra column (col 7 = G) — match col 5 width
    ci6 = ws_in.colinfo_map.get(6)
    ws_out.column_dimensions["G"].width = max((ci6.width/256.0) if ci6 else 10, 10)

    matched   = 0
    not_found = []
    in_batch  = False
    in_summary = False
    batch_idx  = -1
    cur_desig_ref = None
    out_row = 0

    for r in range(nrows):
        if r in dup_rows:
            continue
        out_row += 1
        row = ws_in.row_values(r)

        # Set row height
        ri = ws_in.rowinfo_map.get(r)
        if ri and ri.height and ri.height > 0:
            ws_out.row_dimensions[out_row].height = ri.height / 20.0

        # Copy all cells from template
        for c0 in range(ncols):
            val, is_date = xlrd_val(wb_in, ws_in, r, c0)
            if val is None: continue
            cell = ws_out.cell(row=out_row, column=c0+1, value=val)
            fmt  = cell_fmt(wb_in, ws_in, r, c0)
            if is_date:
                cell.number_format = FMT_DATE
            elif fmt:
                cell.number_format = fmt

        col2_str = str(row[2]).strip()
        col0_str = str(row[0]).strip()

        # Detect state
        if col2_str == 'PartnerID':
            in_batch = True; in_summary = False
            batch_idx += 1
            cur_desig_ref = batch_desig_list[batch_idx] if batch_idx < len(batch_desig_list) else {}
            continue

        if col0_str == 'Total Batch Amount':
            in_batch = False; continue

        if col0_str == 'For Finance':
            in_summary = True; in_batch = False; continue

        # Batch data row
        if in_batch:
            pid = clean_pid(row[0])
            if pid is None:
                in_batch = False; continue
            donor_type = ''; rm = ''
            if pid in lookup:
                donor_type = lookup[pid]['TYPE']
                rm         = lookup[pid]['RM']
                matched   += 1
            else:
                not_found.append(pid)
            ws_out.cell(row=out_row, column=5, value=donor_type)
            ws_out.cell(row=out_row, column=6, value=rm)
            continue

        # Summary header
        if in_summary and col0_str == 'Acount Fund':
            ws_out.cell(row=out_row, column=5, value='Mass')
            ws_out.cell(row=out_row, column=6, value='Middle')
            ws_out.cell(row=out_row, column=7, value='Major')
            continue

        # Summary data rows
        if in_summary:
            non_empty = [v for v in row if v not in ('', None)]
            if not non_empty:
                in_summary = False; continue
            if len(non_empty) == 1 and isinstance(row[0], (int, float)):
                in_summary = False; continue
            designation = str(row[1]).strip() if row[1] not in ('', None) else ''
            if designation and cur_desig_ref:
                ta     = cur_desig_ref.get(designation, {})
                mass   = ta.get('Mass',   0)
                middle = ta.get('Middle', 0)
                major  = ta.get('Major',  0)
                for col_idx, v in [(5, mass), (6, middle), (7, major)]:
                    if v > 0:
                        c = ws_out.cell(row=out_row, column=col_idx, value=v)
                        c.number_format = FMT_RP_SUM
                    else:
                        ws_out.cell(row=out_row, column=col_idx, value='-')

    output = BytesIO()
    wb_out.save(output)
    return output.getvalue(), matched, list(set(not_found)), len(dup_rows)

if sponsor_file and batch_file:
    lookup     = load_sponsor_lookup(sponsor_file)
    file_bytes = batch_file.read()
    result_bytes, matched, not_found, dup_count = process(file_bytes, lookup)

    c1, c2, c3 = st.columns(3)
    c1.metric("Matched",          matched)
    c2.metric("Duplikat dibuang", dup_count)
    c3.metric("Tidak Ditemukan",  len(not_found))

    if not_found:
        with st.expander("PartnerID tidak ditemukan di Data Sponsor"):
            st.write(sorted(not_found))

    st.success("Selesai!")
    st.download_button("Download Hasil", result_bytes, "Crystal_Report_FILLED.xlsx")
else:
    st.info("Silakan upload kedua file untuk memulai.")
